"""

"""
import serial, serial.tools.list_ports
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.scrolledtext as tkst
from tkinter import messagebox
from tkinter import OptionMenu
from threading import Thread
from tkinter.filedialog import askopenfilename
import openpyxl


class SerialThreadClass:
    def __init__(self):
        self._running = False
        self.openedport = None

    def terminate(self):
        self._running = False

    def isThreadRunning(self):
        return self._running

    def write(self,data):
        if self.openedport != None:
            self.openedport.write(data)
        else:
            print('Port is not opened!!!')
            messagebox.showinfo('Error', 'Port is not opened!!!')

    def run(self, port):
        self._running = True
        self.openedport = port
        while self._running:
            try:
                line = self.openedport.readline()
                if (line):
                    print(line)
                    # insert text in a scrolledtext  
                    if gui:
                        gui.updateScrollText(line)           
            except:
                pass            
 
class pedIdWrGui:
    def __init__(self, master):
        self.master = master
        self.OpenedPort = None
        self.master.title("Pedestal ID Writer")
        self.master.resizable(1, 1)
        #master.geometry('300x200')
      
        self.setup()

    def setup(self):
        # Get the COM port
        Ports = serial.tools.list_ports.comports()

        port_var = tk.StringVar(self.master)
        port_var.set(Ports[0])  # initial value

        #Menu Bar
        menu = tk.Menu(self.master)
        self.master.config(menu=menu)
        file = tk.Menu(menu)
        file.add_command(label='debug none', command=lambda t='debug none\r': self.sendcmd(t))     
        file.add_command(label='debug hi', command=lambda t='debug hi\r': self.sendcmd(t))   
        file.add_command(label='dev info', command=lambda t='dev info\r': self.sendcmd(t))    
        file.add_command(label='net info', command=lambda t='net info\r': self.sendcmd(t))
        file.add_command(label='reset', command=lambda t='system reset\r': self.sendcmd(t))
        
        menu.add_cascade(label='Debug Command', menu=file)

        # Create a OptionMenu
        option = OptionMenu(self.master, port_var, command=self.dropdown_get, *Ports)
        #option.config(font=('Helvetica', 12))
        option.grid(column=0, row=0)

        # Create buttons 
        self.xlbtn = tk.Button(text='Open Excel')
        self.xlbtn.grid(column=1, row=0)
        self.xlbtn.config(command=lambda: xl.openFileDialog())

        # Create a scrolledtext
        self.scrt = tkst.ScrolledText(self.master, width=80, height=10, wrap=tk.WORD)
        self.scrt.grid(column=0, row=1, columnspan=2)
        #self.scrt.pack()
        #self.scrt.focus_set()                                          

        # Create buttons 
        self.xlbtn2 = tk.Button(text='ID Write')
        self.xlbtn2.grid(column=2, row=0)
        self.xlbtn2.config(command=lambda: self.idwrite())

        # Create a scrolledtext
        self.scrt2 = tkst.ScrolledText(self.master, width=30, height=10, wrap=tk.WORD)
        self.scrt2.grid(column=2, row=1)
        #self.scrt.pack()
        self.scrt2.focus_set()
    
    def idwrite(self):
        sid, gid, did = xl.getID()

        text = 'S:' + str(sid) + ' G:' + str(gid) + ' ID:' + str(did) + '\r\n'
        #self.scrt2.delete("1.0",tk.END)        
        self.scrt2.insert(tk.INSERT,text)

    def dropdown_get(self, sel_port):
        TargetPort = str(sel_port).split(" ")[0]
        print("[INFO] Use {}...".format(TargetPort))
    
        if (ser.isThreadRunning()):
            print('thread is Run, stop')
            self.OpenedPort.close()
            ser.terminate()

        # Open the COM port
        try:
            self.OpenedPort = serial.Serial(TargetPort, 460800, timeout=0)
        except Exception as ex:
            messagebox.showinfo('Error', ex)
            return
        else:
            if(self.OpenedPort.isOpen() == True):
                self.OpenedPort.close()
            self.OpenedPort.open()

        SerPrtThr = Thread(target=ser.run, args=(self.OpenedPort,))
        SerPrtThr.start()    

    def sendcmd(self, data):
        if ser:
            cmd = data.encode('utf-8')
            ser.write(cmd)

    def updateScrollText(self, data):
        self.scrt.insert(tk.INSERT, data)        
        self.scrt.see(tk.END)

    def refreshXlButtonTxt(self, txtstr):
        self.xlbtn.config(text=txtstr)
        
class excelClass:
    def __init__(self):
        self.initDir = "c:/work/python/"
        self.SiteId = None
        self.GroupID = None
        self.DeviceId = None
        self.excelWB = None
        self.excelName = None

    #This is where we lauch the file manager bar.
    def openFileDialog(self):
        name = askopenfilename(initialdir=self.initDir,
                               filetypes =(("Excel File", "*.xlsx"),("All Files","*.*")),
                               title = "Choose a file."
                              )
        
        if name == '':
            return

        print('Selected File Name is', name)
        
        self.excelName = name

        gui.refreshXlButtonTxt(name)

        # 엑셀파일 열기
        #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
        wb = openpyxl.load_workbook(name, data_only=True)
        self.excelWB = wb
       
    def searchID(self):
        if self.excelWB == None:
            messagebox.showinfo('Error', 'File을 선택후 사용 하세요.')
            return

        wb = self.excelWB
        # 현재 Active Sheet 얻기
        ws = wb.active
        ws = wb['Sheet1']
        #ws = wb['Sheet1'] #시트 이름으로 불러오기
        #print(ws['A1'].value) #셀 주소로 값 출력
        #get_cells = ws['A1':'D2'] #지정한 셀 가져오기
        #ws.columns #모든 열 단위
        #ws.columns #모든 행 단위
        #ws.cell(row=row_index, column=5).value = sum # cell 수정 또는 쓰기
        #wb.save("score2.xlsx") # 엑셀 파일 저장
        #wb.close()        

        isIdFound = False

        for r in ws.rows:
            row_index = r[0].row  # 행 인덱스
            if r[0].value != 'used id':
                isIdFound = True
                self.SiteId = r[1].value
                self.GroupID = r[2].value
                self.DeviceId = r[3].value
                r[0].value = 'used id'
                break
        
        if isIdFound == False :
            messagebox.showinfo('Error', '사용치 않은 device id가 없습니다.')
        else:
            print ('Row is',row_index, 'SiteId:',self.SiteId, 'GroupID:',self.GroupID, 'Device ID:',self.DeviceId)
 
    def getID(self):
        self.searchID()
        return self.SiteId, self.GroupID, self.DeviceId
        
    def terminate(self):
        if self.excelWB != None:
            self.excelWB.save(self.excelName)
            self.excelWB.close()

        #Using try in case user types in unknown file or closes without choosing a file.
        #try:
        #    with open(name,'r') as self.UseFile:
        #        print(self.UseFile.read())
        #except:
        #    print("No file exists")


if __name__ == "__main__":
    root = tk.Tk()
    gui = pedIdWrGui(root)
    ser = SerialThreadClass()
    xl = excelClass()

    root.mainloop()
    ser.terminate()
    xl.terminate()


