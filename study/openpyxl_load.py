"""
파이썬에서 엑셀 데이타를 핸들링하기 위해서는 openpyxl, xlrd, xlrw 등의 외부 패키지를 설치해서 사용한다.
여기서는 openpyxl 패키지를 아래와 같이 설치하여 사용한다.

pip install openpyxl
"""

import openpyxl
 
# 엑셀파일 열기
#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
wb = openpyxl.load_workbook('score.xlsx', data_only=True)
 
# 현재 Active Sheet 얻기
ws = wb.active
# ws = wb.get_sheet_by_name("Sheet1")

#시트 이름으로 불러오기
#ws = wb['Sheet1']

#셀 주소로 값 출력
#print(ws['A1'].value)

#지정한 셀 가져오기
#get_cells = ws['A1':'D2']

#모든 열 단위
#ws.columns

#모든 행 단위
#ws.columns

# 국영수 점수를 읽기
for r in ws.rows:
    row_index = r[0].row   # 행 인덱스
    kor = r[1].value
    eng = r[2].value
    math = r[3].value
    sum = kor + eng + math
 
    # 합계 쓰기
    ws.cell(row=row_index, column=5).value = sum
 
    print(kor, eng, math, sum)
 
# 엑셀 파일 저장
wb.save("score2.xlsx")
wb.close()